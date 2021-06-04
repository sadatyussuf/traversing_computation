import openpyxl
from openpyxl.styles import PatternFill
import math

# lOADING THE EXCEL WORKBOOK
wb = openpyxl.load_workbook('traverse.xlsx')
sheet = wb['Sheet1']
sheet.title = 'Original'


# Input from User
numberOfSides = 5
givenBearing = 0
# numberOfSides = 4
# givenBearing = (140+(11/60)+(40/3600))

givenEasting = 1000
givenNorthing = 1000



pageNum = numberOfSides+3
sheet[f'A{pageNum}'].value = 'Total'




# ------------------------------------------------
# Computing for the sum total of the Angles in Cell B
count = 0
for i in range(2,2+numberOfSides):
    angles = sheet.cell(row=i,column=2).value
    angle = eval(angles[1:])
    count += angle

sheet[f'B{pageNum}'].value = count
# sheet['B8'].value = count
# ---------------xxxxxxxxx-----------------------




# ------------------------------------------------
# Adjustment of angular error
sumObservedAngles = count 
sumActualAngles = 180 * (numberOfSides - 2)
totalError =-(sumObservedAngles - sumActualAngles) 

sheet[f'D{pageNum}'].value =totalError

adjPerAngle  = totalError/numberOfSides
# -------xxxxxxxxx-------------------------------------



# ----------------------------------------------------------------
# Computing for the Total Length in Cell C
sheet.cell(row=1, column=3).value = 'Distance'
count = 0
for i in range(2,2+numberOfSides):
    distance = sheet.cell(row=i, column=3).value
    count += distance

sheet[f'C{pageNum}'].value = count
# -------------xxxxxxxxxxx-----------------------------------------



# --------------------OLd Way Of Adjusting Error--------------------------------
#  Miscole to correct the included angle
# Misclose = 3 * (10 / 3600) * math.sqrt(numberOfSides)

# if Misclose < ((1 / 60) + (50 / 3600)):
#     Misclose = (1 / 60)
# else:
#     Misclose= Misclose

# adjPerAngle = -(Misclose/numberOfSides)
# ----------------------xxxxxxxxxxxx------------------------------------------



# -------------------------------------------------------------
# Computing for the Adjusted misclose per Angle in Cell D
sheet.cell(row=1, column=4).value = 'Correction'
for i in range(2,2+numberOfSides):
    sheet.cell(row=i, column=4).value = adjPerAngle
# -------------------xxxxxxx--------------------------------------




# ----------------------------------------------------------------------
# Computing for the Adjusted Angles in Cell E
sheet.cell(row=1, column=5).value = 'Corrected Angles'
count = 0
for i in range(2,2+numberOfSides):
    angles = sheet.cell(row=i, column=2).value
    sheet.cell(row=i,column=5).value = eval(angles[1:]) + float(sheet.cell(row=i,column=4).value)
    count += sheet.cell(row=i,column=5).value

sheet[f'E{pageNum}'].value = math.ceil(count)
# sheet['E8'].value = round(count)
# ------------------------xxxxxxxxxx----------------------------------------




# ----------------------------------------------------------------
# Computing for the Bearings in Cell F
sheet.cell(row=1, column=6).value = 'W.C.B'
sheet.cell(row=2, column=6).value = givenBearing
for i in range(2,1+numberOfSides):
    correctedAngles = sheet.cell(row=i+1, column=5).value
    if correctedAngles == None:
        correctedAngles=0

    
    WCB = float(sheet.cell(row=i, column=6).value) + 180 + correctedAngles
    # Check if the Bearing is greater than 360, subtract 360 if that's the case
    if WCB > 360:
        WCB = WCB-360
    sheet.cell(row=i+1,column=6).value =WCB

# Check for Bearings
check  = float(sheet.cell(row=pageNum-2,column=6).value) + 180 + float(sheet.cell(row=2, column=5).value)
if check > 360:
    check =check-360
sheet[f'F{pageNum-1}'].fill = PatternFill(bgColor='71FF33')
sheet[f'F{pageNum-1}'].value = check

# -------------xxxxxxxx-----------------------------------------




# --------------------------------------------------------------
# Computing for the Departure in Cell G
sheet.cell(row=1, column=7).value = 'Departure'
for i in range(2,2+numberOfSides):
    distanceOfLine = sheet.cell(row=i, column=3).value
    bearingOfLine = sheet.cell(row=i, column=6).value
    bearingsInRadians = math.radians(bearingOfLine)
    departureOfLine = distanceOfLine * math.sin(bearingsInRadians)
    # Populate column G with the values of departureOfLine
    sheet.cell(row=i,column=7).value = round(departureOfLine,2)


# Computing for the Total Departure in Cell G
count = 0
for i in range(2,2+numberOfSides):
    totalDeparture = sheet.cell(row=i, column=7).value
    count += totalDeparture
sheet[f'G{pageNum}'].value = count
# -----------------------xxxxxxxx-------------------------------





# --------------------------------------------------------------
# Computing for the Latitude in Cell H
sheet.cell(row=1, column=8).value = 'Latitude'
for i in range(2,2+numberOfSides):
    distanceOfLine = sheet.cell(row=i, column=3).value
    bearingOfLine = sheet.cell(row=i, column=6).value
    bearingsInRadians = math.radians(bearingOfLine)
    latitudeOfLine = distanceOfLine * math.cos(bearingsInRadians)
    # Populate column G with the values of departureOfLine
    sheet.cell(row=i,column=8).value = round(latitudeOfLine,2)


# Computing for the Total Latitude in Cell H
count = 0
for i in range(2,2+numberOfSides):
    totalLatitude = sheet.cell(row=i, column=8).value
    count += totalLatitude
sheet[f'H{pageNum}'].value = count
# -----------------------xxxxxxxx-------------------------------




errLat = sheet[f'H{pageNum}'].value
errDep = sheet[f'G{pageNum}'].value

# round(math.sqrt((errDep**2)+(errLat**2)),3)
linearMisclosure = math.sqrt((errDep**2)+(errLat**2))
totalDistance = sheet[f'C{pageNum}'].value

# totalDistance/linearMisclosure
fractionalMilosure = totalDistance/linearMisclosure
sheet[f'B{pageNum+1}'].value = f' 1 / {math.ceil(fractionalMilosure)}'
# Distance of a Line / Total Distance
# def perDistance():
#     pass
# perDistance = 5

# print(f'{totalDistance} / {linearMisclosure} = {fractionalMilosure}')
# print(linearMisclosure)

# --------------------------------------------------------------
# Computing for the Latitude in Cell I
sheet.cell(row=1, column=9).value = 'errorDeparture'
for i in range(2,2+numberOfSides):
    perDistance = sheet.cell(row=i, column=3).value
    distancePerTotalDistance = perDistance/totalDistance
    errorPerLine = distancePerTotalDistance * -sheet[f'G{pageNum}'].value 

    sheet.cell(row=i,column=9).value = errorPerLine

# Computing for the Total errorDeparture in Cell I
count = 0
for i in range(2,2+numberOfSides):
    totalErrorDept = sheet.cell(row=i, column=9).value
    count += totalErrorDept
sheet[f'I{pageNum}'].value = count
# -----------------------xxxxxxxx-------------------------------




# --------------------------------------------------------------
# Computing for the Latitude in Cell J
sheet.cell(row=1, column=10).value = 'errorLatitude'
for i in range(2,2+numberOfSides):
    perDistance = sheet.cell(row=i, column=3).value
    distancePerTotalDistance = perDistance/totalDistance
    errorPerLine = distancePerTotalDistance * -sheet[f'H{pageNum}'].value 

    sheet.cell(row=i,column=10).value = errorPerLine

# Computing for the Total errorLatitude in Cell J
count = 0
for i in range(2,2+numberOfSides):
    totalErrorLat = sheet.cell(row=i, column=10).value
    count +=  totalErrorLat
sheet[f'J{pageNum}'].value = count
# -----------------------xxxxxxxx-------------------------------
wb.save('test1.xlsx')